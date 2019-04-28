#!/usr/bin/env python3
# -*- coding:utf-8 -*-
# @Time     : 2019/4/24 2:09 PM
# @Author   : dnie
# @Email    : niedi761@pingan.com.cn
# @File     : doExcel.py
# @Software : PyCharm


from openpyxl import load_workbook
from homework_0423.common.readconfig import ReadConfig


class DoExcel:
    def __init__(self,file_path,sheet_name):
        self.file_path=file_path
        self.sheet_name=sheet_name
        self.wb=load_workbook(self.file_path)
        self.sheet=self.wb[self.sheet_name]

    def get_init_data(self):
        wb=load_workbook(self.file_path)
        sheet=wb["init"]
        return sheet.cell(1,2).value

    def update_init_data(self,value):
        self.sheet=self.wb["init"]
        self.sheet.cell(1,2).value=value
        self.wb.save(self.file_path)


    def getData(self):
        mobile=self.get_init_data()
        listdata=[]
        for i in range(2,self.sheet.max_row+1):
            subdata={}
            subdata["case_id"]=self.sheet.cell(i,1).value
            subdata["title"]=self.sheet.cell(i,2).value
            subdata["url"]=ReadConfig().getValue("data","baseurl")+self.sheet.cell(i,3).value
            if self.sheet.cell(i,4).value.find("${reg_mobile}") != -1:
                new_data=self.sheet.cell(i,4).value.replace("${reg_mobile}",str(mobile))
            else:
                new_data=self.sheet.cell(i,4).value
            subdata["data"] = new_data
            subdata["method"]=self.sheet.cell(i,5).value
            subdata["expected"]=self.sheet.cell(i,6).value
            subdata["check_sql"]=self.sheet.cell(i,9).value
            listdata.append(subdata)
        return listdata


    def writeData(self,row,column,value):
        self.wb = load_workbook(self.file_path)
        self.sheet = self.wb[self.sheet_name]
        self.sheet.cell(row,column).value=value
        self.wb.save(self.file_path)
        self.wb.close()